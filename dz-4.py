# Импортируем библиотеку pandas, collections
import pandas, collections
# Импортируем функцию из библиотеки openpyxl
from openpyxl import load_workbook


# Функция, возвращающая список ранжированных по популярности объектов (браузеров или товаров)
    # аргументы функции:
        # listing - список всех ранжируемых объектов;
        # key_name - ключ в словаре excel_data_dict для ранжируемых объектов;
        # quantity - количество объектов в возвращаемом списке pop_list по популярности
def popular(listing, key_name, quantity):
    pop_listing = collections.Counter(listing).most_common(quantity) # список популярных объектов
    pop_list=[]
    for element in pop_listing:
        m=[0]*13 # список, который будет содержать название объекта, общее количество за год и количество объектов за каждый месяц
        m[0]=element[1]
        for i in range(len(excel_data_dict)):
            if element[0] in excel_data_dict[i][key_name]:
                jear_month_day = str(excel_data_dict[i]['Дата посещения'])
                m[int(jear_month_day[5:7])] +=1
        pop_list.append([element[0], m[0], m[1], m[2], m[3], m[4], m[5], m[6], m[7], m[8], m[9], m[10], m[11], m[12]])
    return pop_list


# Читаем файл ексель и результат передаем в переменную excel_data
# Переменная excel_data имеет тип <class 'pandas.core.frame.DataFrame'>
excel_data = pandas.read_excel('logs.xlsx', sheet_name='log')

# Преобразуем переменную excel_data в список словарей с помощью метода to_dict()
# Результат передаем в переменную excel_data_dict
excel_data_dict = excel_data.to_dict(orient='records')

browsers = [] # список всех браузеров
items = [] # список всех купленных товаров
man_items = [] # список товаров, купленных мужчинами
woman_items = [] # список товаров, купленных женщинами
for element in excel_data_dict:
    browsers.append(element['Браузер'])
    s = element['Купленные товары'].split(',') # из строки товаров одного пользователя получаем список товаров 
    for k in s:
        items.append(k.strip()) # из начала и конца строки с названием каждого товара удаляем пробелы и добавляем в общий список товаров 
        if element['Пол']=='м':
            man_items.append(k.strip()) # товары, купленные мужчинами добавляем в список man_items
        else:
            woman_items.append(k.strip()) # товары, купленные женщинами добавляем в список woman_items      


browser_list = popular(browsers, 'Браузер', 7) # список ранжированных по популярности браузеров
item_list = popular(items, 'Купленные товары', 7) # список ранжированных по популярности товаров
item_man = collections.Counter(man_items).most_common() # ранжированный по популярности список товаров, купленных мужчинами
pop_item_man = item_man[0][0] # Самый популярный товар среди мужчин
nopop_item_man = item_man[-1][0] # Самый нeвостребованный товар среди мужчин
item_woman = collections.Counter(woman_items).most_common() # ранжированный по популярности список товаров, купленных женщинами
pop_item_woman = item_woman[0][0] # Самый популярный товар среди женщин
nopop_item_woman = item_woman[-1][0] # Cамый невостребованный товар среди женщин


# открываем файл ексель'report.xlsx' и записываем в него полученные результаты 
wb = load_workbook(filename='report.xlsx')
sheet = wb['Лист1']
for i in range(7):
    for j in range(14):
        sheet.cell(row=i+5, column=j+1).value = browser_list[i][j]
        sheet.cell(row=i+19, column=j+1).value = item_list[i][j]
sheet["B31"] = pop_item_man
sheet["B32"] = pop_item_woman
sheet["B33"] = nopop_item_man
sheet["B34"] = nopop_item_woman
wb.save(filename='report.xlsx')
